import os
import sys
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
OWNER        = "tijanaminic1"
REPO         = "pid2-pr4"
OUTPUT_FILE  = "pr_comments.xlsx"
# ─────────────────────────────────────────────────────────────────────────────

if not GITHUB_TOKEN:
    print("Error: GITHUB_TOKEN environment variable is not set.", file=sys.stderr)
    print("  export GITHUB_TOKEN=<your_token>", file=sys.stderr)
    sys.exit(1)

BASE_URL = "https://api.github.com"
HEADERS  = {
    "Authorization": f"Bearer {GITHUB_TOKEN}",
    "Accept": "application/vnd.github+json",
    "X-GitHub-Api-Version": "2022-11-28",
}


def get_all_pages(url: str, params: dict = None) -> list:
    """Fetch all pages from a paginated GitHub API endpoint."""
    params = {**(params or {}), "per_page": 100}
    items = []
    while url:
        resp = requests.get(url, headers=HEADERS, params=params)
        resp.raise_for_status()
        items.extend(resp.json())
        params = {}  # params only needed on first request; next URL already has them
        url = resp.links.get("next", {}).get("url")
    return items


def fetch_all_prs() -> list:
    url = f"{BASE_URL}/repos/{OWNER}/{REPO}/pulls"
    return get_all_pages(url, {"state": "all"})


def pr_state(pr: dict) -> str:
    if pr.get("merged_at"):
        return "merged"
    return pr["state"]


def fetch_comments(pr_number: int, pr_title: str) -> list[dict]:
    """Return all three comment types for a single PR as a flat list of rows."""
    rows = []
    repo_name = f"{OWNER}/{REPO}"

    # 1. PR-level (issue) comments
    url = f"{BASE_URL}/repos/{OWNER}/{REPO}/issues/{pr_number}/comments"
    for c in get_all_pages(url):
        rows.append({
            "body": c.get("body") or "",
            "repo": repo_name,
            "pr_number": pr_number,
            "pr_title": pr_title,
            "comment_type": "pr-level",
            "file_path": "",
            "line": "",
            "diff_hunk": "",
        })

    # 2. Inline review (pull request) comments
    url = f"{BASE_URL}/repos/{OWNER}/{REPO}/pulls/{pr_number}/comments"
    for c in get_all_pages(url):
        rows.append({
            "body": c.get("body") or "",
            "repo": repo_name,
            "pr_number": pr_number,
            "pr_title": pr_title,
            "comment_type": "inline",
            "file_path": c.get("path") or "",
            "line": c.get("line") or c.get("original_line") or "",
            "diff_hunk": c.get("diff_hunk") or "",
        })

    # 3. Review submissions
    url = f"{BASE_URL}/repos/{OWNER}/{REPO}/pulls/{pr_number}/reviews"
    for r in get_all_pages(url):
        body = (r.get("body") or "").strip()
        if body:
            rows.append({
                "body": body,
                "repo": repo_name,
                "pr_number": pr_number,
                "pr_title": pr_title,
                "comment_type": "review",
                "file_path": "",
                "line": "",
                "diff_hunk": "",
            })

    return rows


def build_excel(all_rows: list[dict]) -> None:
    headers = [
        "Comment Body",         # A
        "Topic 1",              # B
        "Topic 2",              # C
        "Comp: Problem ID",     # D
        "Comp: Suggestion",     # E
        "Comp: Question",       # F
        "Comp: Praise",         # G
        "Comp: Ref to Spec",    # H
        "Comp: Inconsistency",  # I
        "Validity",             # J
        "Actionability",        # K
        "Reasoning",            # L
        "Required vs Optional", # M
        "Perceived Authorship", # N
        "Participant ID",       # O
        "PR Number",            # P
        "Repo Name",            # Q
        "Comment Location",     # R
    ]

    header_font    = Font(bold=True)
    header_fill    = PatternFill("solid", fgColor="D9E1F2")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    top_alignment  = Alignment(vertical="top")

    existing_pairs: set[tuple] = set()
    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        for sheet_row in ws.iter_rows(min_row=2, values_only=True):
            pid_val = sheet_row[14]  # column O — Participant ID
            pr_val  = sheet_row[15]  # column P — PR Number
            if pid_val is not None and pr_val is not None:
                existing_pairs.add((str(pid_val), str(pr_val)))
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "PR Comments"
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = top_alignment

    skipped = 0
    for row in all_rows:
        if row["comment_type"] == "inline" and row["diff_hunk"]:
            location = row["diff_hunk"]
        else:
            location = row["comment_type"]

        # Parse participant ID and PR number from repo name
        # e.g. "tijanaminic1/pid2-pr4" -> participant_id=2, pr_number=4
        repo_short = row["repo"].split("/")[-1]  # "pid2-pr4"
        try:
            parts = repo_short.split("-")
            participant_id = parts[0].replace("pid", "")
            pr_num = parts[1].replace("pr", "")
        except (IndexError, ValueError):
            participant_id = ""
            pr_num = ""

        if (str(participant_id), str(pr_num)) in existing_pairs:
            skipped += 1
            continue

        ws.append([
            row["body"],        # A — Comment Body
            "",                 # B — Topic 1
            "",                 # C — Topic 2
            "",                 # D — Comp: Problem identification
            "",                 # E — Comp: Suggestion
            "",                 # F — Comp: Question
            "",                 # G — Comp: Praise
            "",                 # H — Comp: Ref to Spec
            "",                 # I — Comp: Inconsistency
            "",                 # J — Validity
            "",                 # K — Actionability
            "",                 # L — Reasoning
            "",                 # M — Required vs Optional
            "",                 # N — Perceived Authorship
            participant_id,     # O — Participant ID (auto-filled)
            pr_num,             # P — PR Number (auto-filled)
            row["repo"],        # Q — Repo Name
            location,           # R — Comment Location
        ])

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            if cell.column in (1, 18):  # Comment Body and Comment Location
                cell.alignment = wrap_alignment
            else:
                cell.alignment = top_alignment

    # Column widths
    ws.column_dimensions["A"].width = 80
    for col_letter in ["B", "C"]:
        ws.column_dimensions[col_letter].width = 25
    for col_letter in ["D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col_letter].width = 18
    for col_letter in ["J", "K", "L", "M", "N"]:
        ws.column_dimensions[col_letter].width = 22
    for col_letter in ["O", "P"]:
        ws.column_dimensions[col_letter].width = 15
    ws.column_dimensions["Q"].width = 30
    ws.column_dimensions["R"].width = 40

    ws.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)
    print(f"Saved {ws.max_row - 1} rows to {OUTPUT_FILE} ({skipped} skipped as duplicates)")


def main():
    print(f"Fetching PRs from {OWNER}/{REPO} ...")
    prs = fetch_all_prs()
    print(f"Found {len(prs)} PRs\n")

    all_rows = []
    for pr in prs:
        number = pr["number"]
        title  = pr["title"]
        state  = pr_state(pr)
        rows   = fetch_comments(number, title)
        print(f"PR #{number} ({state}): {title!r} — {len(rows)} comment(s)")
        all_rows.extend(rows)

    print(f"\nTotal comments scraped: {len(all_rows)}")
    build_excel(all_rows)


if __name__ == "__main__":
    main()
